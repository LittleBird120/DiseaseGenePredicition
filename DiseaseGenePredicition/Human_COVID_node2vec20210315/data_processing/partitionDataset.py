import json
import math
import os
import networkx as nx
import openpyxl
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split

def split(data,ratio_train):
    train,test = train_test_split(data,test_size=1-ratio_train)
    return train,test

#数据集划分
def partition(parent_path,host_filename):
    train_path = os.path.join(parent_path,r'data\crossValidation\firstExperiments\train.txt')
    test_path = os.path.join(parent_path,r'data\crossValidation\firstExperiments\test.txt')
    with open(host_filename, 'r') as node_fr:
        nodeslines = node_fr.readlines()
        hostDict = json.loads(nodeslines[0])

    hostList = []
    for i in range(len(hostDict)):
        hostList.append(hostDict[i][1])

    raw_dataset = np.array(hostList)
    train_dataset, test_dataset = split(raw_dataset, 0.8)
    print(train_dataset)

    np.savetxt(train_path, train_dataset, fmt='%s')
    np.savetxt(test_path, test_dataset, fmt='%s')

    return train_dataset,test_dataset

#所有数据集
def allData(parent_path,host_filename):
    train_path = os.path.join(parent_path,r'data\crossValidation\secondExperiments\train.txt')
    with open(host_filename, 'r') as node_fr:
        nodeslines = node_fr.readlines()
        hostDict = json.loads(nodeslines[0])

    hostList = []
    for i in range(len(hostDict)):
        hostList.append(hostDict[i][1])

    train_dataset = np.array(hostList)
    print(train_dataset)

    np.savetxt(train_path, train_dataset, fmt='%s')

    return train_dataset

#对训练集进行网络构建
def getTrainNetwork(train_dataset,file_edges,file_hostsPro):
    with open(file_edges, 'r') as edge_fr:
        edgeslines = edge_fr.readlines()
        edgesList = json.loads(edgeslines[0])

    with open(file_hostsPro, 'r') as node_fr:
        nodeslines = node_fr.readlines()
        nodesDict = json.loads(nodeslines[0])

    G = nx.Graph()
    for node in nodesDict.keys():
        G.add_node(node)
    for edge in edgesList:
        G.add_edge(edge[0], edge[1])
    print("图形已构成")

    adj = G._adj
    trainHost = []
    trainEdgesList = []
    for everyTrainHost in train_dataset:
        trainHost.append(everyTrainHost)
        adj_host = list(adj[everyTrainHost].keys())
        trainHost.extend(adj_host)
    trainHost = list(set(trainHost))

    for i in range(len(trainHost)):
        for j in range(i+1, len(trainHost)):
            if G.has_edge(trainHost[i], trainHost[j]):
                trainEdgesList.append([trainHost[i], trainHost[j]])

    # trainHostlist = []
    # for i in range(len(trainEdgesList)):
    #     trainHostlist.extend(trainEdgesList[i])
    # trainHostlist = list(set(trainHostlist))

    trainHostDict = {}
    for node in trainHost:
        trainHostDict[node] = nodesDict[node]

    result_host_path = parent_path + r'\data\crossValidation\secondExperiments\train_host.json'
    with open(result_host_path, 'w') as fw:
        json.dump(trainHostDict, fw)

    result_virus_host_path = parent_path + r'\data\crossValidation\secondExperiments\train_virus_host.json'
    with open(result_virus_host_path, 'w') as fw:
        json.dump(trainEdgesList, fw)

    return trainHostDict,trainEdgesList

#构建每个蛋白质的注释term及其祖先term的字典
def build_goDict(file_Go,trainHostDict):
    ontologylist = ["BP", "MF", "CC"]
    all = []
    for ontology in ontologylist:
        dict = {}
        go_df = pd.read_excel(file_Go, sheet_name=ontology)  # 返回一个DataFrame的对象，这个是pandas的一个数据结构
        protein = go_df["protein"]
        ancestor = go_df["all"]
        protein = np.array(protein)
        ancestor = np.array(ancestor)

        for i in range(len(protein)):
            if protein[i] in trainHostDict.keys():
                dict[protein[i]] = ancestor[i]
        all.append(dict)

    pro = list(trainHostDict.keys())
    return all, pro

#获取节点的边
def getEdges(edgesList):
    firArr = []
    secArr = []
    for v in edgesList:
        firArr.append(v[0])
        secArr.append(v[1])
    firArr = np.array(firArr)
    secArr = np.array(secArr)

    return firArr, secArr

#求不同本体下基于go的蛋白质对之间的相似度
def everyOntology_sim(all_dict,fir, sec):
    ontology_list = []
    for dict in all_dict:
        inter_dict = {}
        for index in range(len(fir)):
            ontology_sim = 0.00
            common_go = 0  # 共有term的个数
            all_go = 0  # 蛋白质对的所有term的个数
            if (fir[index] in dict.keys()) and (sec[index] in dict.keys()):
                fir_go = dict[fir[index]]
                sec_go = dict[sec[index]]
                if fir_go == "NULL" or sec_go == "NULL" or isinstance(fir_go, float) or isinstance(sec_go, float):
                    ontology_sim = 0.00
                else:
                    # fir_go_list = fir_go.split("|")
                    # sec_go_list = sec_go.split("|")
                    # common_list = [x for x in fir_go_list if x in sec_go_list]  # term对的并集
                    # common_go = len(common_list)
                    # all_list = list(set(fir_go_list).union(set(sec_go_list)))  # term对的交集
                    # all_go = len(all_list)
                    # ontology_sim = round(common_go / all_go, 2)

                    fir_go_list = fir_go.split("|")
                    sec_go_list = sec_go.split("|")
                    common_list = [x for x in fir_go_list if x in sec_go_list]  # term对的并集
                    common_go = len(common_list)
                    ontology_sim = round(math.sqrt(math.pow(common_go,2)/(len(fir_go_list)*len(sec_go_list))),2)
            else:
                ontology_sim = 0.00
            interaction = fir[index] + "|" + sec[index]
            inter_dict[interaction] = ontology_sim
        ontology_list.append(inter_dict)
    return ontology_list

#综合三个本体的相似度
def ComprehensiveSim(ontology_list):
    sim = 0.00
    dict = {}
    for key,valueBP in ontology_list[0].items():
        valueMF = ontology_list[1][key]
        valueCC = ontology_list[2][key]
        sim = round((valueBP + valueMF + valueCC)/3,2)
        dict[key] = sim
    return dict

#保存基于go的相似度
def save_goSim(dict, pro):
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    i = 1
    outws.cell(i, 1).value = "protein"
    outws.cell(i, 2).value = "neghbor_protein"
    outws.cell(i, 3).value = "sim"
    for key,value in dict.items():
        proteinList = key.split("|")
        source = pro.index(proteinList[0])
        target = pro.index(proteinList[1])
        proteinList[0] = str(source) + "_" + proteinList[0]
        proteinList[1] = str(target) + "_" + proteinList[1]
        proteinList.append(value)
        for j in range(len(proteinList)):
            outws.cell(i + 1, j+1).value = proteinList[j]
        i = i + 1

    file_path = "../data/crossValidation/secondExperiments/train_COVID_AllHuman_GoSim.xls"
    outwb.save(file_path)  # 一定要记得保存

if __name__ == '__main__':
    ROOT_DIR = os.path.dirname(os.path.abspath('__file__'))
    parent_path = os.path.dirname(ROOT_DIR)
    file_Go = "../data/COVID_HumanOntology.xls"
    file_edges = "../data/uploads/virus_host.json"
    file_hostsPro = "../data/uploads/hostpro.json"
    file_hosts = "../data/uploads/PartialHost.json"
    # train_dataset, test_dataset = partition(parent_path,file_hosts)
    train_dataset = allData(parent_path, file_hosts)
    trainHostDict,trainEdgesList = getTrainNetwork(train_dataset,file_edges,file_hostsPro)
    all_dict, pro = build_goDict(file_Go,trainHostDict)
    fir, sec = getEdges(trainEdgesList)
    ontology_list = everyOntology_sim(all_dict, fir, sec)
    dict = ComprehensiveSim(ontology_list)
    save_goSim(dict, pro)